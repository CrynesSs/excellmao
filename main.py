from pathlib import Path
from copy import copy
import configparser

import pandas as pd
from openpyxl import load_workbook


config = configparser.ConfigParser()
config.read("./config/init.ini")

# --- Paths ---
csv_path = Path(config["PATHS"]["csv_path"])
mapping_excel_path = Path(config["PATHS"]["mapping_excel_path"])
target_excel_path = Path(config["PATHS"]["target_excel_path"])
output_path = Path(config["PATHS"]["output_path"])

# --- Settings ---
update_existing = config.getboolean("SETTINGS", "update_existing_file")
# If update existing -> output path == input path
if update_existing:
    output_path = target_excel_path

# --- CSV config ---
csv_key_col = config["CSV"]["key_col"]
csv_value_1_col = config["CSV"]["value_1_col"]
csv_value_2_col = config["CSV"]["value_2_col"]

# --- Mapping Excel config ---
mapping_sheet_name = config["MAPPING_EXCEL"]["sheet_name"]
mapping_key_col = config["MAPPING_EXCEL"]["key_col"]
mapping_worksheet_col = config["MAPPING_EXCEL"]["worksheet_col"]
mapping_place_1_col = config["MAPPING_EXCEL"]["place_1_col"]
mapping_place_2_col = config["MAPPING_EXCEL"]["place_2_col"]

# --- Style config ---
font_color = config["STYLE"].get("font_color", "FFFF0000")

if not update_existing and target_excel_path.resolve() == output_path.resolve():
    raise AssertionError(
        "Target input and output are the same, while update_existing_file=false"
    )

save_path = target_excel_path if update_existing else output_path


# --- Load CSV write data ---
df_csv = pd.read_csv(csv_path, sep=",")

required_csv_cols = [csv_key_col, csv_value_1_col, csv_value_2_col]
missing_csv_cols = [col for col in required_csv_cols if col not in df_csv.columns]

if missing_csv_cols:
    raise ValueError(f"Missing CSV columns: {missing_csv_cols}")

df_csv[csv_key_col] = df_csv[csv_key_col].astype(str).str.strip()

if df_csv[csv_key_col].duplicated().any():
    duplicates = df_csv.loc[df_csv[csv_key_col].duplicated(), csv_key_col].tolist()
    raise ValueError(f"Duplicate IDs found in CSV: {duplicates}")

csv_lookup = df_csv.set_index(csv_key_col)[
    [csv_value_1_col, csv_value_2_col]
].to_dict(orient="index")


# --- Load mapping workbook: ID -> worksheet, Place1, Place2 ---
mapping_wb = load_workbook(mapping_excel_path, data_only=True)
mapping_ws = mapping_wb[mapping_sheet_name]

header_row = 1

mapping_headers = {
    cell.value: cell.column
    for cell in mapping_ws[header_row]
    if cell.value is not None
}

required_mapping_cols = [
    mapping_key_col,
    mapping_worksheet_col,
    mapping_place_1_col,
    mapping_place_2_col,
]

missing_mapping_cols = [
    col for col in required_mapping_cols
    if col not in mapping_headers
]

if missing_mapping_cols:
    raise ValueError(f"Missing mapping Excel columns: {missing_mapping_cols}")

mapping_key_col_idx = mapping_headers[mapping_key_col]
mapping_worksheet_col_idx = mapping_headers[mapping_worksheet_col]
mapping_place_1_col_idx = mapping_headers[mapping_place_1_col]
mapping_place_2_col_idx = mapping_headers[mapping_place_2_col]

mapping_lookup = {}

for row in range(header_row + 1, mapping_ws.max_row + 1):
    mapping_id = mapping_ws.cell(row=row, column=mapping_key_col_idx).value

    if mapping_id is None:
        continue

    mapping_id = str(mapping_id).strip()

    if mapping_id in mapping_lookup:
        raise ValueError(f"Duplicate ID found in mapping workbook: {mapping_id}")

    worksheet_name = mapping_ws.cell(row=row, column=mapping_worksheet_col_idx).value
    place_1 = mapping_ws.cell(row=row, column=mapping_place_1_col_idx).value
    place_2 = mapping_ws.cell(row=row, column=mapping_place_2_col_idx).value

    if worksheet_name is None or str(worksheet_name).strip() == "":
        raise ValueError("Missing target worksheet name")
    else:
        worksheet_name = str(worksheet_name).strip()

    mapping_lookup[mapping_id] = {
        "worksheet": worksheet_name,
        "place_1": place_1,
        "place_2": place_2,
    }


# --- Load target workbook, preserving macros if it is xlsm ---
keep_vba = target_excel_path.suffix.lower() == ".xlsm"

target_wb = load_workbook(
    target_excel_path,
    keep_vba=keep_vba,
    rich_text=True,
)


def write_red_value(workbook, worksheet_name, cell_address, value, color):
    """
    Writes value to workbook[worksheet_name][cell_address] in red font.
    Preserves existing font styling except color.
    """
    if worksheet_name is None or str(worksheet_name).strip() == "":
        raise ValueError("Missing target worksheet name")

    worksheet_name = str(worksheet_name).strip()

    if worksheet_name not in workbook.sheetnames:
        raise ValueError(f"Target worksheet does not exist: {worksheet_name}")

    if cell_address is None or str(cell_address).strip() == "":
        return False

    cell_address = str(cell_address).strip()

    ws = workbook[worksheet_name]
    target_cell = ws[cell_address]

    target_cell.value = value

    new_font = copy(target_cell.font)
    new_font.color = color
    target_cell.font = new_font

    return True


matched = 0
csv_without_mapping = 0
written_values = 0

for csv_id, csv_values in csv_lookup.items():
    if csv_id not in mapping_lookup:
        csv_without_mapping += 1
        continue

    mapping = mapping_lookup[csv_id]

    worksheet_name = mapping["worksheet"]
    place_1 = mapping["place_1"]
    place_2 = mapping["place_2"]

    value_1 = csv_values[csv_value_1_col]
    value_2 = csv_values[csv_value_2_col]

    if write_red_value(target_wb, worksheet_name, place_1, value_1, font_color):
        written_values += 1

    if write_red_value(target_wb, worksheet_name, place_2, value_2, font_color):
        written_values += 1

    matched += 1


target_wb.save(save_path)

print("Done.")
print(f"Matched IDs: {matched}")
print(f"CSV IDs without mapping: {csv_without_mapping}")
print(f"Written values: {written_values}")
print(f"Saved to: {save_path}")